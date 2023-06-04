# Программа возвращает кадастровые номера земельных участков по адресу, убирая снятые с учета
# В результате создает таблицу в указанной папке и открывает ее.
# Исходная таблица должна быть .xlsx и содержать 4 колонки:
# id	Тип		НаименУлицы    Дом
# ---------------------------------
# 12    улица   Скоростная		12

import os
import random
import openpyxl
import time
import rosreestr_online

start = time.time() ## точка отсчета времени

file_name = "результат" + str(random.randint(1, 10000)) + ".xlsx"

wb = openpyxl.load_workbook('src.xlsx')
sheet = wb.active

table_output = openpyxl.Workbook()
sheet_table_output = table_output.active

# Создаем заголовки в результирующей таблице
sheet_table_output.append(('id', 'Тип', 'НаимУлицы', 'Дом', "КадастровыйНомерЗУ"))
max_row = sheet.max_row

for i in range(2, max_row):
	id = sheet[i][0].value
	type_street = sheet[i][1].value
	ul = sheet[i][2].value
	number = str(sheet[i][3].value)
	cadnums = rosreestr_online.getFullCadNumParcel(type_street, ul, number)  # Получаем список кад номеров
	cadnums = '; '.join(str(cadnum) for cadnum in cadnums) # Преобразуем его в строку, элементы разделяем "; "

	# Заполняем строку данными
	object = []
	object.append(id)
	object.append(type_street)
	object.append(ul)
	object.append(number)
	object.append(cadnums)

	print(object)

	# Добавляем в результирующую таблицу и сохраняем ее
	sheet_table_output.append(object)
	table_output.save(file_name)

end = time.time() - start #  время работы программы
os.startfile(file_name) # Открываем результирующий файл
print(end) # вывод времени
